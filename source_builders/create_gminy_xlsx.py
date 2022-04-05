## create gminy.xlsx with boroughs names only in A col (from C col)
# borough -> gmina

from openpyxl import load_workbook
from openpyxl import Workbook

def filter_cond_city_name(value):
    if value is None:
        return False
    else:
        return True

def clean_garbage(city_name):
    if city_name.endswith(" m"):
        city_name = (city_name[:-2]).strip()
    if "\n" in city_name or "\r" in city_name:
        city_name = city_name.replace("\n", "").replace("\r", "")
    if city_name.count("-") > 1:
        city_name = city_name.replace("-", "", city_name.count("-") - 1)  # - vs -
    return city_name

wb = load_workbook(filename ='wykaz_stan_na_1.1.2019.xlsx')
sheet_ranges = wb['Arkusz1']
unq_set = set()
iterator = 0
for e in sheet_ranges['C']:
    if filter_cond_city_name(e.value):
        new_city_name = clean_garbage(e.value)
        unq_set.add(new_city_name)
        iterator += 1
    if iterator % 100 == 0:
        print(f"{e.value}.")
if iterator < len(unq_set):
    raise Exception("Fatal error")
unq_set = sorted(tuple(unq_set))

my_wb = Workbook()
new_xlsx_file = "gminy.xlsx"
my_ws = my_wb.active
my_ws.title="Gminy"


idx = 1
for ue in unq_set:
    my_ws[f'A{idx}'] = ue
    idx += 1

my_wb.save(filename=new_xlsx_file)
#print(sheet_ranges['D18'].value)